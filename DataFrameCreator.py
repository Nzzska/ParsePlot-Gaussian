import os
import re
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class ExcitedStatesParser():
    def __init__(self, filesDirectory, numberOfExcitedStates):
        self.numberOfExcitedStates = numberOfExcitedStates
        if filesDirectory == '':
            self.filesDirectory = os.path.dirname(os.path.realpath(__file__))
        else:
            self.filesDirectory = filesDirectory
        self.findLogs()
        self.sortLogs()
        self.normalizeEnergies()
        self.groupJumpsData()

    def findLogs(self):
        self.allLogs = list()
        angle = 0
        statesInfo =   list()

        for filename in os.listdir(self.filesDirectory):
            if filename.endswith('.log'):

                splitedFilename = filename.split('.')
                splitedFilename = splitedFilename[0]
                splitedFilename = splitedFilename.split('n')

                #Creating array of turn Angles used for Jobs
                angle = float(splitedFilename[1])

                #Returns information about energies, and electron jumps in list [energies, jumps, OscillatorStrengths]
                #FOR SINGLE FILE
                statesInfo = self.addExcitedStates(filename)
                statesInfo.append(angle)

                self.allLogs.append(statesInfo)

    def addExcitedStates(self, filename):
        allJumps = list()
        with open(os.path.join(self.filesDirectory, filename)) as fhandle:
            energies =            [0] * (self.numberOfExcitedStates + 1)
            OscillatorStrengths = [0] * (self.numberOfExcitedStates + 1)
            waveLengths         = [0] * (self.numberOfExcitedStates + 1)
            #WORKAROUNDO MASYVAI
            jumps1 = []
            jumps2 = []
            jumps3 = []
            #############
            jumps = []
            for line in fhandle:
                for i in range(0, self.numberOfExcitedStates + 1 , 1):
                    #Takes in i-state number and line to parse, returns dict {energy, OscillatorStrength} if information found.
                    energiesAndOscillators = self.nthExcitedState(i, line)
                    if energiesAndOscillators != None :
                        activeState = i
                #appending energy to appropriate list
                        energies[i] = energiesAndOscillators['energy']
                #repeating action for Oscillator Strengths
                        OscillatorStrengths[i] = energiesAndOscillators['oscillatorStrength']
                        waveLengths[i] = energiesAndOscillators['wavelength']
                #Searching for information about jumps
                jump1 = re.findall("[0-9]{3}\s->[0-9]{3}", line)
                jump2 = re.findall("[0-9]{3}\s<-[0-9]{3}", line)
                if (jump1 != []) | (jump2 != []):
                    words = line.split()
                    jump = words[0] + words[1]
                    intensity = words[2]
                    jumpData = {jump : intensity}
                    #Hard codintas workaroundas ! TODO!
                    if activeState == 1:
                        jumps1.append(jumpData)
                    elif activeState == 2:
                        jumps2.append(jumpData)
                    elif activeState == 3:
                        jumps3.append(jumpData)
        jumps = [jumps1, jumps2, jumps3]
                    #Hard codintas workaroundas baigiasi
        return [
            energies, 
            jumps, 
            OscillatorStrengths,
            waveLengths
        ]


    def nthExcitedState(self, excitedState, line):
        try:
            if excitedState == 0:
                if line.startswith(' SCF Done:  E(RB3LYP) ='):
                    line.strip()
                    words = line.split()
                    energy = float(words[4]) * (-1) * 27.211386245988
                    return {
                        'energy': energy,
                        'oscillatorStrength': 'NaN',
                        'wavelength': 'NaN'
                        }
            else:
                if line.startswith(' Excited State   {}'.format(str(excitedState))):
                    line.strip()
                    words = line.split()
                    energy = float(words[4])
                    osc = words[8].split('=')
                    oscillatorStrength = float(osc[1])
                    wavelength = float(words[6])
                    return {
                        'energy': energy,
                        'oscillatorStrength': oscillatorStrength,
                        'wavelength': wavelength
                    }
        except:
            print('excited state {} doesn\'t exist'.format(str(excitedState)))
            print(line)

    def sortLogs(self):
        anglesVariable = len(self.allLogs[0]) - 1
        for i in range(0, len(self.allLogs), 1):
            for j in range(0, len(self.allLogs) - 1, 1):
                if self.allLogs[j][anglesVariable] > self.allLogs[j + 1][anglesVariable]:
                    tempLog = self.allLogs[j + 1]
                    self.allLogs[j + 1] = self.allLogs[j]
                    self.allLogs[j] = tempLog

    def normalizeEnergies(self):
        groundStates = list()
        for log in self.allLogs:
            groundStates.append(log[0][0])
        baseLine = min(groundStates)
        for log in self.allLogs:
            log[0][0] -= baseLine
            for i in range(1, len(log[0]), 1):
                log[0][i] += log[0][0]

    def groupJumpsData(self):
        for log in self.allLogs:
            jumpsMultiple = []
            for jumpList in log[1]:
                #jumpsInfo yra list'as kur suoliu info sudame i 
                jumpSingle = {}
                for entry in jumpList:
                    copy = entry.copy()
                    jumpSingle.update(copy)
                jumpsMultiple.append(jumpSingle)
            log.append(jumpsMultiple)


class excitedStatesDataFrame():
    def __init__(self, excitedStatesData):
        self.dataFrame = pd.DataFrame(excitedStatesData)

    def exportAllData(self):
        with pd.ExcelWriter('Visi duomenys.xlsx', engine = 'openpyxl') as writer:
            self.dataFrame.to_excel(writer, sheet_name='visi duomenys')

    def exportJumpsData(self):
        row = 1
        with pd.ExcelWriter('excitedStatesInfo.xlsx', engine = 'openpyxl') as writer:
            for i , DataFrameColumns in self.dataFrame.iterrows():
                #sitie iki col= 10 eina. jumps data
                jumpsDataFrame = pd.DataFrame(DataFrameColumns[5])
                jumpsDataFrame.to_excel(writer, sheet_name = 'Informacija apie šuolius', startrow = row, startcol = 0)
                row += 5
        row = 2
        wb = openpyxl.load_workbook('excitedStatesInfo.xlsx')
        ws = wb.active
        ws.cell(row = 1, column = 1).value = 'ExctState:'
        ws['A1'].font = Font(bold = True)
        ws.cell(row = 1, column = 11).value = 'Turn Angle:'
        ws['K1'].font = Font(bold = True)
        ws.cell(row = 1, column = 13).value = 'Oscillator Strengths:'
        ws['M1'].font = Font(bold = True)
        ws.cell(row = 1, column = 15).value = 'Delta energies (eV):'
        ws['O1'].font = Font(bold = True)
        ws.merge_cells(start_row = 1, start_column = 2, end_row = 1, end_column = 8)
        ws.cell(row = 1, column =2).value = 'Electron jumps info'
        ws.cell(row = 1, column =2).alignment = Alignment(horizontal = "center", vertical = "center")
        ws.cell(row = 1, column =2).font = Font(bold = True)
        for i , DataFrameColumns in self.dataFrame.iterrows():
            ws.cell(row = row, column = 11).value = DataFrameColumns[4]
            ws.cell(row = row, column = 1).fill = PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
            ws.cell(row = row + 1, column = 1).value = 1
            ws.cell(row = row + 2, column = 1).value = 2
            ws.cell(row = row + 3, column = 1).value = 3
            tempRow = 1
            for oscillator in DataFrameColumns[2]:
                if oscillator != 'NaN':
                    ws.cell(row = row + tempRow, column = 13).value = oscillator
                    tempRow += 1
            tempRow = 0
            for energy in DataFrameColumns[0]:
                    ws.cell(row = row + tempRow, column = 15).value = energy
                    tempRow += 1
            wb.save('excitedStatesInfo.xlsx')
            row += 5

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value




####

haha = ExcitedStatesParser('', 3)
hihi = excitedStatesDataFrame(haha.allLogs)
hihi.exportJumpsData()
hihi.exportAllData()